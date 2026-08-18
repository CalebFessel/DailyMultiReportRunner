"""
Report output layer: Excel formatting, append workbooks, email, housekeeping.

Lifted verbatim from the original runner so formatting, de-duplication,
retention and delivery behave exactly as before. Nothing here knows where the
data came from, which is the whole point -- swapping ODBC for the Traumasoft
API changes only how the DataFrames are produced, not what happens to them.
"""

import os
import re
import ssl
import sys
import logging
import smtplib
from pathlib import Path
from email.message import EmailMessage
from datetime import datetime, timedelta

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# =============================
# CONFIG
# =============================
OUTPUT_DIR = os.getenv("OUTPUT_DIR", "Reports")
APPEND_DIR = os.path.join(OUTPUT_DIR, "Append")
RETENTION_DAYS = int(os.getenv("RETENTION_DAYS", "14"))
APPEND_RETENTION_DAYS = int(os.getenv("APPEND_RETENTION_DAYS", "730"))
EXCEL_TABLE_STYLE = os.getenv("EXCEL_TABLE_STYLE", "TableStyleMedium9")


# =============================
# LOGGING SETUP
# =============================
def setup_logging(output_dir: str, run_date_str: str) -> str:
    log_dir = os.path.join(output_dir, "logs")
    Path(log_dir).mkdir(parents=True, exist_ok=True)
    log_path = os.path.join(log_dir, f"DailyReports_{run_date_str}.log")

    # Reconfigure handlers for this run
    for h in list(logging.getLogger().handlers):
        logging.getLogger().removeHandler(h)

    logging.basicConfig(
        level=logging.INFO,
        encoding="utf-8",
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.FileHandler(log_path), logging.StreamHandler(sys.stdout)],
    )
    return log_path


# =============================

# =============================
# EXCEL HELPERS (TABLES)
# =============================
def _sanitize_table_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", str(name))
    if not cleaned or not re.match(r"^[A-Za-z_]", cleaned):
        cleaned = f"T_{cleaned}"
    return cleaned[:200]


def _autofit_columns(ws, max_width=60):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        best = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            best = max(best, len(str(v)))
        if best > 0:
            ws.column_dimensions[letter].width = min(max_width, best + 2)


def add_table_to_sheet(ws, table_name: str, style_name: str = EXCEL_TABLE_STYLE):
    if ws.max_row < 2 or ws.max_column < 1:
        return False
    last_col = get_column_letter(ws.max_column)
    last_row = ws.max_row
    ref = f"A1:{last_col}{last_row}"
    tname = _sanitize_table_name(table_name)
    table = Table(displayName=tname, ref=ref)
    style = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    ws.freeze_panes = "A2"
    _autofit_columns(ws)
    return True


def write_df_sheet_with_table(writer, df: pd.DataFrame, sheet_name: str, table_name: str):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.book[sheet_name]
    add_table_to_sheet(ws, table_name=table_name, style_name=EXCEL_TABLE_STYLE)


# =============================

# =============================
# APPEND HELPERS
# =============================
def _safe_sheet_name(s: str) -> str:
    # Excel sheet names max 31, no []:*?/\
    s = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(s))
    return s[:31] if len(s) > 31 else s


def _append_to_workbook_xlsx(
    append_path: str,
    sheet_name: str,
    df_new: pd.DataFrame,
    dedupe_keys=None,
    snapshot_date_value=None,
    snapshot_col: str = "snapshot_date",
):
    """
    Append df_new to append_path/sheet_name, de-duping on keys if provided.
    - If append file doesn't exist -> create
    - If sheet doesn't exist -> create
    - Adds snapshot_date column if snapshot_date_value provided and column missing
    - Prunes old rows if snapshot_date or work_date column exists (APPEND_RETENTION_DAYS)
    """
    Path(os.path.dirname(append_path)).mkdir(parents=True, exist_ok=True)

    df = df_new.copy()

    if snapshot_date_value is not None and snapshot_col not in df.columns:
        df.insert(0, snapshot_col, snapshot_date_value)

    sheet_name = _safe_sheet_name(sheet_name)

    existing = None
    if os.path.exists(append_path):
        try:
            existing = pd.read_excel(append_path, sheet_name=sheet_name, engine="openpyxl")
        except ValueError:
            existing = None
        except Exception:
            existing = None

    if existing is None:
        combined = df
    else:
        combined = pd.concat([existing, df], ignore_index=True, sort=False)

    if dedupe_keys:
        present_keys = [k for k in dedupe_keys if k in combined.columns]
        if present_keys:
            combined = combined.drop_duplicates(subset=present_keys, keep="last")

    cutoff_date = datetime.now().date() - timedelta(days=APPEND_RETENTION_DAYS)

    if snapshot_col in combined.columns:
        sd = pd.to_datetime(combined[snapshot_col], errors="coerce").dt.date
        combined = combined.loc[(sd.isna()) | (sd >= cutoff_date)].copy()
    elif "work_date" in combined.columns:
        wd = pd.to_datetime(combined["work_date"], errors="coerce").dt.date
        combined = combined.loc[(wd.isna()) | (wd >= cutoff_date)].copy()

    if os.path.exists(append_path):
        with pd.ExcelWriter(
            append_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace",
        ) as writer:
            write_df_sheet_with_table(
                writer,
                combined,
                sheet_name=sheet_name,
                table_name=f"APPEND_{sheet_name}",
            )
    else:
        with pd.ExcelWriter(append_path, engine="openpyxl", mode="w") as writer:
            write_df_sheet_with_table(
                writer,
                combined,
                sheet_name=sheet_name,
                table_name=f"APPEND_{sheet_name}",
            )


# =============================

# =============================
# EMAIL / HOUSEKEEPING
# =============================
def send_email(subject, body, recipients, attachments=None):
    smtp_server = os.getenv("SMTP_SERVER", "smtp.example.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASS") or os.getenv("SMTP_PASSWORD")
    smtp_from = os.getenv("SMTP_FROM", smtp_user)
    ehlo_host = os.getenv("SMTP_EHLO_HOST", "alerts.example.com")

    if not smtp_user or not smtp_password:
        raise RuntimeError("Missing SMTP_USER/SMTP_PASS (or SMTP_PASSWORD) in environment variables")

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = smtp_from
    msg["To"] = ", ".join(recipients)
    msg.set_content(body)

    for path in (attachments or []):
        with open(path, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=os.path.basename(path),
            )

    with smtplib.SMTP(smtp_server, smtp_port, timeout=30, local_hostname=ehlo_host) as server:
        code, resp = server.ehlo()
        logging.info(f"EHLO response: {code} {resp!r}")
        logging.info(f"ESMTP features: {server.esmtp_features}")
        if code != 250:
            raise RuntimeError(f"EHLO failed ({code}): {resp!r}")
        if not server.has_extn("starttls"):
            raise RuntimeError(f"STARTTLS not offered. Features: {server.esmtp_features}")

        server.starttls(context=ssl.create_default_context())
        server.ehlo()
        server.login(smtp_user, smtp_password)
        server.send_message(msg)


def cleanup_old_files():
    cutoff = datetime.now() - timedelta(days=RETENTION_DAYS)
    for f in Path(OUTPUT_DIR).glob("*.xlsx"):
        if datetime.fromtimestamp(f.stat().st_mtime) < cutoff:
            try:
                f.unlink()
                logging.info(f"Deleted old file: {f}")
            except Exception as e:
                logging.warning(f"Failed to delete {f}: {e}")


